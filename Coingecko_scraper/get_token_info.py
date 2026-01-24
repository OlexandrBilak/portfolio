import os
import time
from datetime import datetime
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook, load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

FILENAME = "coingecko.xlsx"
start_time = datetime.now()


def get_driver():
    options = Options()
    # options.add_argument("--headless=new")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--lang=en-US")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver


def create_table():
    if os.path.exists(FILENAME):
        return

    wb = Workbook()
    ws = wb.active
    ws.append([
        "created_at_utc",
        "case_source",
        "token_name",
        "token_symbol",
        "volume_24h",
        "volatility_24h",
        "token_chain",
        "token_contract",
        "token_link",
        "company_name",
        "company_website",
        "company_socials",
        "company_twitter",
        "company_telegram",
        "company_linkedin"
        "listings_count",
        "CEX_count",
        "DEX_count",
    ])
    wb.save(FILENAME)


def process_tokens_from_excel(limit=300):
    wb = load_workbook(FILENAME)
    ws = wb.active
    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    driver = get_driver()
    wait = WebDriverWait(driver, 25)
    processed = 0

    try:
        for row in range(2, ws.max_row + 1):
            if processed >= limit:
                break

            token_link = ws.cell(row=row, column=headers["token_link"]).value
            token_contract_exists = ws.cell(row=row, column=headers["token_contract"]).value

            if not token_link or token_contract_exists:
                continue

            print(f"[PROCESS] {processed + 1} → {token_link}")
            driver.get(token_link)

            wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            time.sleep(2)

                        # --- SCROLL TO LOAD ALL CONTENT ---
            last_height = driver.execute_script("return document.body.scrollHeight")
            while True:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(1)
                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height

            soup_all = BeautifulSoup(driver.page_source, "lxml")
            soup = soup_all.find('div', class_='tw-relative 2lg:tw-mb-6 tw-grid tw-grid-cols-1 tw-divide-y tw-divide-gray-200 dark:tw-divide-moon-700 [&>*:last-child]:!tw-border-b')

            # --- BASIC DATA ---
            try:
                token_contract = soup.find(
                    "div",
                    class_="tw-flex-1 tw-flex tw-items-center tw-gap-x-2"
                ).text
            except:
                token_contract = "none"

            company_website = ""
            company_name = ""
            try:
                website_block = soup.find("div", string=lambda x: x and "Вебсайт" in x).find_parent("div")
                links = website_block.find_all("a")
                if links:
                    company_website = links[0].get("href", "")
                    company_name = links[0].text.strip()
            except:
                pass


            company_socials = []
            company_twitter = ""
            company_telegram = ""
            company_linkedin = ""

            try:
                community_block = soup.find(
                    "div",
                    string=lambda x: x and "Спільнота" in x
                ).find_parent("div")

                for a in community_block.find_all("a"):
                    href = a.get("href")
                    if not href:
                        continue

                    company_socials.append(href)

                    if "twitter.com" in href or "x.com" in href:
                        company_twitter = href
                    elif "t.me" in href:
                        company_telegram = href
                    elif "linkedin.com" in href:
                        company_linkedin = href
            except:
                pass

            company_socials_str = ", ".join(company_socials)

            
            # --- MARKET DATA VIA SELENIUM ---
            listings_count = 0
            cex_count = 0
            dex_count = 0

            try:
                # клік на вкладку Ринки
                markets_tab = wait.until(
                    EC.element_to_be_clickable((By.ID, "tab-markets"))
                )
                markets_tab.click()
                time.sleep(3)  # чекаємо, поки підвантажаться рядки

                # скролимо вниз, щоб підвантажити всі рядки, якщо потрібно
                last_height = driver.execute_script("return document.body.scrollHeight")
                while True:
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(1)
                    new_height = driver.execute_script("return document.body.scrollHeight")
                    if new_height == last_height:
                        break
                    last_height = new_height

                market_rows = driver.find_elements(By.CSS_SELECTOR, "div[data-coin-show-target='markets'] tbody > tr")
                listings_count = len(market_rows)

                for row_elem in market_rows:
                    try:
                        market_type = row_elem.find_element(By.CSS_SELECTOR, "td:nth-child(3) span").text.strip()
                        if market_type == "CEX":
                            cex_count += 1
                        elif market_type == "DEX":
                            dex_count += 1
                    except:
                        pass

            except Exception as e:
                print("[MARKETS ERROR]", e)

            # --- WRITE TO EXCEL ---
            ws.cell(row=row, column=headers["created_at_utc"]).value = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row=row, column=headers["token_contract"]).value = token_contract
            ws.cell(row=row, column=headers["company_website"]).value = company_website
            ws.cell(row=row, column=headers["company_name"]).value = company_name
            ws.cell(row=row, column=headers["company_socials"]).value = company_socials_str
            ws.cell(row=row, column=headers["company_twitter"]).value = company_twitter
            ws.cell(row=row, column=headers["company_telegram"]).value = company_telegram
            ws.cell(row=row, column=headers["company_linkedin"]).value = company_linkedin
            ws.cell(row=row, column=headers["listings_count"]).value = listings_count
            ws.cell(row=row, column=headers["CEX_count"]).value = cex_count
            ws.cell(row=row, column=headers["DEX_count"]).value = dex_count

            print(f'{company_name} - {company_socials} - {cex_count} - {dex_count}')
            wb.save(FILENAME)
            processed += 1
            time.sleep(2)

    finally:
        driver.quit()

    print(f"[DONE] processed {processed} tokens")


if __name__ == "__main__":
    create_table()
    process_tokens_from_excel(limit=3)
    print(f"Час роботи: {datetime.now() - start_time}")
