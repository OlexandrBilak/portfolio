from openpyxl import load_workbook

FILENAME = "coinmarketcap_listings_lt_5.xlsx"  # ← твій файл


def extract_token_chain_from_tags():
    wb = load_workbook(FILENAME)
    ws = wb.active

    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    col_tags = headers["tags"]
    col_chain = headers["token_chain"]

    updated = 0
    skipped = 0

    for row in range(2, ws.max_row + 1):
        tags_val = ws.cell(row=row, column=col_tags).value

        if not tags_val:
            skipped += 1
            continue

        ecosystems = []

        for tag in tags_val.split(","):
            tag = tag.strip()

            if tag.lower().endswith("ecosystem"):
                # прибираємо слово "Ecosystem"
                chain = tag[:-len("Ecosystem")].strip()
                ecosystems.append(chain)

        if ecosystems:
            ws.cell(row=row, column=col_chain).value = ", ".join(ecosystems)
            updated += 1
        else:
            skipped += 1

    wb.save(FILENAME)
    print(f"[DONE] updated={updated}, skipped={skipped}")


if __name__ == "__main__":
    extract_token_chain_from_tags()
