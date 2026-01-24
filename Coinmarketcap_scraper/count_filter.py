from openpyxl import load_workbook

FILENAME = "coinmarketcap_listings_lt_5.xlsx"  # ← зміни на свій файл


def recalc_dex_from_listings():
    wb = load_workbook(FILENAME)
    ws = wb.active

    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    col_listings = headers["listings_count"]
    col_cex = headers["CEX_count"]
    col_dex = headers["DEX_count"]

    updated = 0
    skipped = 0

    for row in range(2, ws.max_row + 1):
        listings = ws.cell(row=row, column=col_listings).value
        cex = ws.cell(row=row, column=col_cex).value

        if listings is None or cex is None:
            skipped += 1
            continue

        try:
            listings = int(listings) - 1
            cex = int(cex)

            dex = listings - cex
            if dex < 0:
                dex = 0

            ws.cell(row=row, column=col_dex).value = dex
            ws.cell(row=row, column=col_listings).value = listings

            updated += 1

        except Exception:
            skipped += 1

    wb.save(FILENAME)

    print(f"[DONE] updated={updated}, skipped={skipped}")


if __name__ == "__main__":
    recalc_dex_from_listings()
