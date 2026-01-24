import time
from openpyxl import load_workbook
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from webdriver_manager.chrome import ChromeDriverManager


FILENAME = "coinmarketcap_listings_lt_5.xlsx"


def get_driver():
    options = Options()
    # options.add_argument("--headless=new")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--lang=en-US")

    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def click_tab(driver, tab_test_id):
    # чекаємо на присутність вкладки
    tab_li = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, f"//li[@data-test='{tab_test_id}']"))
    )

    # якщо вкладка вже активна, не клікаємо
    if "Tab_selected__" in tab_li.get_attribute("class"):
        return

    tab_label = tab_li.find_element(By.XPATH, ".//div[@data-role='label']")

    driver.execute_script("""
        arguments[0].scrollIntoView({block: 'center'});
        arguments[0].click();
    """, tab_label)

    # чекаємо поки вкладка стане активною
    WebDriverWait(driver, 10).until(
        lambda d: "Tab_selected__" in tab_li.get_attribute("class")
    )

    time.sleep(0.5)


def count_rows_from_table(driver):
    soup = BeautifulSoup(driver.page_source, "lxml")
    empty = soup.select_one(".EmptyTableContent_empty-table-content__no-results")
    if empty:
        return 0
    rows = soup.select("table.cmc-table tbody tr")
    return len(rows)


def process_excel(limit=50):
    wb = load_workbook(FILENAME)
    ws = wb.active

    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    driver = get_driver()
    processed = 0
    skipped = 0

    try:
        for row in range(2, ws.max_row + 1):
            if processed >= limit:
                break

            token_link = ws.cell(row=row, column=headers["token_link"]).value
            cex_val = ws.cell(row=row, column=headers["CEX_count"]).value
            dex_val = ws.cell(row=row, column=headers["DEX_count"]).value

            if not token_link or (cex_val is not None and dex_val is not None):
                skipped += 1
                continue

            print(f"[PROCESS] row {row} → {token_link}")
            driver.get(token_link)

            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            time.sleep(2)

            # --- CEX ---
            try:
                click_tab(driver, "cex")
                cex_count = count_rows_from_table(driver)
            except Exception as e:
                print("CEX error:", e)
                cex_count = 0

            # --- DEX ---
            try:
                click_tab(driver, "dex")
                dex_count = count_rows_from_table(driver)
            except Exception as e:
                print("DEX error:", e)
                dex_count = 0

            ws.cell(row=row, column=headers["CEX_count"]).value = cex_count
            ws.cell(row=row, column=headers["DEX_count"]).value = dex_count

            wb.save(FILENAME)
            processed += 1
            print(f"[OK] row {row} | CEX={cex_count} | DEX={dex_count}")
            time.sleep(1)

    finally:
        driver.quit()

    print(f"[DONE] processed={processed} skipped={skipped}")


if __name__ == "__main__":
    process_excel(limit=3)
