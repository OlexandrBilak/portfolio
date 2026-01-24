from openpyxl import load_workbook, Workbook

SOURCE_FILE = "coinmarketcap.xlsx"
OUTPUT_FILE = "coinmarketcap_listings_lt_5.xlsx"

LISTINGS_COLUMN_NAME = "listings_count"


def filter_by_listings_count():
    wb = load_workbook(SOURCE_FILE)
    ws = wb.active

    # зчитуємо заголовки
    headers = [cell.value for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers)}

    if LISTINGS_COLUMN_NAME not in header_index:
        raise ValueError(f"Column '{LISTINGS_COLUMN_NAME}' not found")

    listings_col_idx = header_index[LISTINGS_COLUMN_NAME] + 1

    # новий файл
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(headers)

    copied = 0

    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=listings_col_idx).value

        try:
            value_int = int(value)
        except (TypeError, ValueError):
            continue

        if value_int <= 5:
            row_data = [
                ws.cell(row=row, column=col).value
                for col in range(1, ws.max_column + 1)
            ]
            new_ws.append(row_data)
            copied += 1

    new_wb.save(OUTPUT_FILE)

    print(f"[DONE] copied {copied} rows to {OUTPUT_FILE}")


if __name__ == "__main__":
    filter_by_listings_count()
