import os
import time
from datetime import datetime
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook, load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

FILENAME = "coinmarketcap.xlsx"
start_time = datetime.now()


def get_driver():
    options = Options()
    # options.add_argument("--headless=new")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--lang=en-US")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver


def create_table():
    if os.path.exists(FILENAME):
        return

    wb = Workbook()
    ws = wb.active
    ws.append([
        "created_at_utc",
        "case_source",
        "token_name",
        "token_symbol",
        "volume_24h",
        "volatility_24h",
        "token_chain",
        "token_contract",
        "token_link",
        "company_name",
        "company_website",
        "company_socials",
        "company_twitter",
        "company_telegram",
        "listings_count",
        "CEX_count",
        "DEX_count",
        "tags"
    ])
    wb.save(FILENAME)


def process_tokens_from_excel(limit=300):
    wb = load_workbook(FILENAME)
    ws = wb.active
    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    driver = get_driver()
    wait = WebDriverWait(driver, 25)
    processed = 0

    try:
        for row in range(2, ws.max_row + 1):
            if processed >= limit:
                break

            token_link = ws.cell(row=row, column=headers["token_link"]).value
            token_contract_exists = ws.cell(row=row, column=headers["token_contract"]).value

            if not token_link or token_contract_exists:
                continue

            print(f"[PROCESS] {processed + 1} → {token_link}")
            driver.get(token_link)

            wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            time.sleep(2)

            soup = BeautifulSoup(driver.page_source, "lxml")

            # --- BASIC DATA ---
            try:
                token_contract = soup.find(
                    "div",
                    class_="BasePopover_base__T5yOf popover-base LongTextDisplay_content-fg__yyqLY LongTextDisplay_content-fg-normal__kPRiv ContractLinks_contract-display-item-content-wrapper__q14bO"
                ).text
            except:
                token_contract = "none"

            try:
                company_website = soup.find(
                    "a",
                    class_="BaseChip_base__lzmlq BaseChip_size-md__A17GX BaseChip_vd__hUjl8 ActionChip_base__sMLzN ActionChip_p0__0CsOR ActionChip_has-interaction__mbyOO"
                ).get("href")
            except:
                company_website = ""

            try:
                company_name = company_website.split(".")[1].split("/")[0]
            except:
                company_name = ""

            company_socials = []
            company_twitter = ""
            company_telegram = ""

            try:
                socials = soup.find_all(
                    "a",
                    class_="BaseChip_base__lzmlq BaseChip_size-md__A17GX BaseChip_vd__hUjl8 BaseChip_only-icon__7QEyH ActionChip_base__sMLzN ActionChip_p0__0CsOR ActionChip_has-interaction__mbyOO SocialLinks_social-action-chip-wrapper__46oKa"
                )

                for s in socials:
                    href = "https:" + s.get("href")
                    company_socials.append(href)
                    if href.startswith("https://twitter.com/"):
                        company_twitter = href
                    elif href.startswith("https://t.me/"):
                        company_telegram = href
            except:
                pass

            company_socials_str = ", ".join(company_socials)

            # --- CLICK MARKETS TAB ---
            listings_count = ""

            try:
                markets_btn = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, '//span[@data-test="section-coin-markets"]')
                    )
                )

                # скрол до кнопки
                driver.execute_script(
                    "arguments[0].scrollIntoView({block: 'center'});",
                    markets_btn
                )
                time.sleep(1)

                # клік
                driver.execute_script("arguments[0].click();", markets_btn)

                wait.until(EC.presence_of_element_located((By.TAG_NAME, "tbody")))
                time.sleep(2)

                soup = BeautifulSoup(driver.page_source, "lxml")
                listings_count = len(
                    soup.find("table", class_="sc-7e3c705d-3 eKbMGC cmc-table").find_all("tr")
                )
            except Exception as e:
                print("[MARKETS ERROR]", e)

            # --- TAGS ---
            try:
                tags = [t.text for t in soup.find_all(
                    "span", class_="sc-65e7f566-0 sc-9ee74f67-1 eQBACe izfTnl"
                )]
                tags_str = ", ".join(tags)
            except:
                tags_str = ""

            # --- WRITE TO EXCEL ---
            ws.cell(row=row, column=headers["created_at_utc"]).value = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row=row, column=headers["token_contract"]).value = token_contract
            ws.cell(row=row, column=headers["company_website"]).value = company_website
            ws.cell(row=row, column=headers["company_name"]).value = company_name
            ws.cell(row=row, column=headers["company_socials"]).value = company_socials_str
            ws.cell(row=row, column=headers["company_twitter"]).value = company_twitter
            ws.cell(row=row, column=headers["company_telegram"]).value = company_telegram
            ws.cell(row=row, column=headers["listings_count"]).value = listings_count
            ws.cell(row=row, column=headers["tags"]).value = tags_str

            wb.save(FILENAME)
            processed += 1
            time.sleep(2)

    finally:
        driver.quit()

    print(f"[DONE] processed {processed} tokens")


if __name__ == "__main__":
    create_table()
    process_tokens_from_excel(limit=3)
    print(f"Час роботи: {datetime.now() - start_time}")
