import openpyxl
import datetime
from grok2 import grok_generate
from validator3 import validate_text, correct_text_by_gpt
import traceback

start_time = datetime.datetime.now()

def log(msg, level="INFO"):
    print(f"[{level}] {msg}")

wb = openpyxl.load_workbook("products.xlsx")
ws = wb.active

HEADER = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

def get_cell(row, name):
    return ws.cell(row=row, column=HEADER[name]).value or ""

def set_cell(row, name, value):
    ws.cell(row=row, column=HEADER[name]).value = value

for row in range(2, ws.max_row + 1):
    try:
        article = get_cell(row, "Артикул")
        name_ru = get_cell(row, "Название (RU)")
        name_ua = get_cell(row, "Название (UA)")
        brand = get_cell(row, "Бренд")
        category = get_cell(row, "Раздел")

        if not name_ru and not name_ua and not brand and not category:
            continue

        for lang in ["RU", "UA"]:
            log(f"▶ Початок обробки товару {article} ({lang})")

            # 1) Початкова генерація Grok
            try:
                log("   → Grok: генерація тексту")
                result = grok_generate(
                    lang=lang,
                    name=name_ru if lang == "RU" else name_ua,
                    brand=brand,
                    category=category
                )
            except Exception as e:
                log(f"- ПОМИЛКА при grok_generate: {e}", "ERROR")
                log(traceback.format_exc())
                continue

            # 2) Перевірка OK/NOT_OK (для логів)
            try:
                validation = validate_text(lang, result)
                if validation["status"] != "OK":
                    log("   → Стара валідація: текст потребує корекції", "WARNING")
            except Exception as e:
                log(f"- ПОМИЛКА при validate_text: {e}", "ERROR")
                log(traceback.format_exc())

            # 3) Виклик нової функції для реальної корекції
            try:
                log("   → ChatGPT: виправлення тексту у строгому форматі")
                corrected_result = correct_text_by_gpt(lang, result)
            except Exception as e:
                log(f"- ПОМИЛКА при correct_text_by_gpt: {e}", "ERROR")
                log(traceback.format_exc())
                continue

            # 4) Збереження виправленого тексту
            set_cell(row, f"HTML title ({lang})", corrected_result["title"])
            set_cell(row, f"META keywords ({lang})", corrected_result["keywords"])
            set_cell(row, f"META description ({lang})", corrected_result["description"])
            set_cell(row, f"h1 заголовок ({lang})", corrected_result["h1"])
            set_cell(row, f"Описание товара ({lang})", corrected_result["content"])

            log(f"✔ Товар {article} ({lang}) оброблено та збережено", "SUCCESS")

    except Exception as e_outer:
        log(f"- ПОМИЛКА при обробці рядка {row}: {e_outer}", "ERROR")
        log(traceback.format_exc())

wb.save("products_ready.xlsx")
log("! Файл оброблено та збережено", "SUCCESS")
print(f"Затрачено час: {datetime.datetime.now() - start_time}")
