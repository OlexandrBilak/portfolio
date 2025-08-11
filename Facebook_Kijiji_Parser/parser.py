import json
import csv
import time
import datetime
import requests
import os
import re
from bs4 import BeautifulSoup
from configparser import ConfigParser
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

if os.path.exists("seen_ads.json"):
    with open("seen_ads.json", "r", encoding="utf-8") as f:
        seen_ads = set(json.load(f))
else:
    seen_ads = set()

def send_to_telegram(message, token, chat_id):
    try:
        requests.post(f"https://api.telegram.org/bot{token}/sendMessage", data={
            "chat_id": chat_id,
            "text": message
        })
    except Exception as e:
        print(f"⚠️ Telegram send error: {e}")

def send_file_to_telegram(file_path, token, chat_id):
    try:
        with open(file_path, "rb") as f:
            requests.post(
                f"https://api.telegram.org/bot{token}/sendDocument",
                data={"chat_id": chat_id},
                files={"document": f}
            )
    except Exception as e:
        print(f"⚠️ Не вдалося надіслати файл у Telegram: {e}")

def extract_year(text):
    match = re.search(r"(19|20)\d{2}", text)
    return int(match.group()) if match else None

def extract_price(text):
    match = re.search(r"(\d[\d\s,]{2,})\s*CAD", text)
    if match:
        return match.group(1).replace(" ", "").replace(",", "")
    return None

def parse_age_to_hours(text):
    if not text:
        return 9999
    text = text.lower()
    match = re.search(r"\d+", text)
    if not match:
        if "just" in text or "now" in text or "today" in text:
            return 0
        return 9999
    value = int(match.group())
    if "min" in text:
        return value // 60
    elif "hour" in text or "hr" in text:
        return value
    elif "day" in text:
        return value * 24
    elif "wk" in text:
        return value * 7 * 24
    elif "mo" in text:
        return value * 30 * 24
    return 9999

def parse_facebook(config):
    print("🔍 Парсимо Facebook...")

    options = Options()
    options.add_argument(f"user-data-dir={config['facebook_profile_path']}")
    options.add_argument(f"profile-directory={config['facebook_profile_name']}")
    # Без headless, щоб бачити браузер
    options.add_experimental_option("detach", True)
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")

    try:
        driver = webdriver.Chrome(options=options)
    except Exception as e:
        print(f"❌ Не вдалося запустити браузер: {e}")
        return []

    query = f"{config['make']} {config['model']}"
    fb_url = (
        f"https://www.facebook.com/marketplace/{config['city_fb']}/search"
        f"?minPrice={config['price_min']}&maxPrice={config['price_max']}&query={query.replace(' ', '%20')}&exact=false"
    )
    if config['facebook_days_filter'] in [1, 7, 30]:
        fb_url += f"&daysSinceListed={config['facebook_days_filter']}"

    print(f"🔗 FB URL: {fb_url}")
    try:
        driver.get(fb_url)
    except Exception as e:
        print(f"❌ Сторінка Facebook не відкрилась: {e}")
        driver.quit()
        return []

    time.sleep(5)
    for _ in range(5):
        driver.execute_script("window.scrollBy(0, document.body.scrollHeight);")
        time.sleep(3)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    driver.quit()

    items = soup.find_all("a", href=True)
    if not items:
        print("⚠️ [Facebook] Жодного <a href> не знайдено! Можливо не завантажилась сторінка?")

    new_ads = []

    for item in items:
        href = item.get("href", "")
        if "/marketplace/item/" in href:
            title = item.text.strip()
            url = "https://www.facebook.com" + href.split("?")[0]
            if url in seen_ads or not title:
                continue
            if config['make'].lower() not in title.lower() or config['model'].lower() not in title.lower():
                continue
            year = extract_year(title)
            if not year or not (config['year_min'] <= year <= config['year_max']):
                continue
            price = extract_price(title)

            new_ads.append({
                "title": title,
                "url": url,
                "price": f"{price} CAD" if price else "—",
                "year": str(year),
                "date_posted": f"Last {config['facebook_days_filter']} days"
            })
            seen_ads.add(url)

    print(f"🌐 Facebook — знайдено {len(new_ads)} нових.")
    return new_ads

def parse_kijiji(config):
    print(f"🌐 Парсимо Kijiji...")
    make_model = f"{config['make']}-{config['model']}".replace(" ", "-")

    city_code_map = {
        "toronto": "l1700273",
        "mississauga": "l1700220",
        "brampton": "l1700221",
        "ottawa": "l1700185",
        "hamilton": "l1700209",
        "london": "l1700214",
        "kitchener": "l1700212",
        "windsor": "l1700223",
        "vancouver": "l1700287",
        "surrey": "l1700291",
        "burnaby": "l1700288",
        "richmond": "l1700289",
        "kelowna": "l1700228",
        "victoria": "l1700173",
        "calgary": "l1700199",
        "edmonton": "l1700203",
        "red deer": "l1700200",
        "saskatoon": "l1700197",
        "regina": "l1700196",
        "winnipeg": "l1700192",
        "halifax": "l1700321",
        "montreal": "l1700081",
        "quebec city": "l1700124",
        "guelph": "l1700242",
        "barrie": "l1700006",
        "sherbrooke": "l1700121",
        "moncton": "l1700001",
        "fredericton": "l1700002",
        "st. john's": "l1700111",
    }

    city_slug = config['city_kijiji'].lower()
    location_id = city_code_map.get(city_slug, "1700273")  # дефолт: Торонто

    kijiji_url = (
        f"https://www.kijiji.ca/b-cars-trucks/{city_slug}/{make_model}/"
        f"{config['year_min']}__{config['year_max']}/k0c174l{location_id}a68"
        f"?price={config['price_min']}__{config['price_max']}&sort=dateDesc&view=list"
    )

    print(f"🔗 Kijiji URL: {kijiji_url}")

    options = Options()
    options.add_argument(f"user-data-dir={config['kijiji_profile_path']}")
    options.add_argument(f"profile-directory={config['kijiji_profile_name']}")
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")

    try:
        driver = webdriver.Chrome(options=options)
    except Exception as e:
        print(f"❌ Не вдалося запустити браузер для Kijiji: {e}")
        return []

    driver.get(kijiji_url)
    time.sleep(6)
    for _ in range(10):
        driver.execute_script("window.scrollBy(0, document.body.scrollHeight);")
        time.sleep(3)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    driver.quit()

    listings = soup.find_all("div", class_="sc-336af988-9")
    new_ads = []

    for i, listing in enumerate(listings):
        try:
            link_tag = listing.find("a", href=True)
            title_tag = listing.find("a", class_="sc-336af988-1")
            price_tag = listing.find("p", {"data-testid": "autos-listing-price"})
            date_tag = listing.find("p", {"data-testid": "listing-date"})

            if not link_tag or not title_tag:
                continue

            url = link_tag["href"].split("?")[0]
            title = title_tag.text.strip()
            price_text = price_tag.text.strip().replace(",", "").replace("$", "") if price_tag else ""
            date_text = date_tag.text.strip() if date_tag else ""

            if url in seen_ads:
                continue
            if not title:
                continue

            price = int(price_text) if price_text.isdigit() else None
            if price is None or not (config['price_min'] <= price <= config['price_max']):
                continue

            year = extract_year(title)
            if not year or not (config['year_min'] <= year <= config['year_max']):
                continue

            hours = parse_age_to_hours(date_text)
            if hours > config['max_post_age_hours']:
                continue

            new_ads.append({
                "title": title,
                "url": url,
                "price": f"{price} CAD",
                "year": str(year),
                "date_posted": date_text
            })
            seen_ads.add(url)

        except Exception as e:
            print(f"⚠️ [{i}] Помилка з оголошенням: {e}")
            continue

    print(f"🌐 Kijiji — знайдено {len(new_ads)} нових.")
    return new_ads




def save_ads(ads, make, model, city_fb, fb_filter_label):
    now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"ads_{now}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Оголошення"

    headers = ["Назва", "Посилання", "Ціна", "Рік", "Дата", "Марка", "Модель", "Місто", "FB фільтр"]
    ws.append(headers)

    # Зробити заголовки жирними
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)

    for ad in ads:
        row = [
            ad.get("title"),
            ad.get("url"),
            ad.get("price"),
            ad.get("year"),
            ad.get("date_posted"),
            make,
            model,
            city_fb,
            fb_filter_label
        ]
        ws.append(row)

    # Зробити посилання клікабельним (гіперпосиланням)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            url = cell.value
            cell.hyperlink = url
            cell.style = "Hyperlink"

    # Автоматична ширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    wb.save(filename)
    return filename

def save_seen():
    with open("seen_ads.json", "w", encoding="utf-8") as f:
        json.dump(list(seen_ads), f, indent=2, ensure_ascii=False)

def validate_text_field(value, field_name):
    if field_name in ["kijiji_profile_path"]:
        return value.strip()
    if not re.match(r"^[a-zA-Z0-9\- \\:/]+$", value):
        raise ValueError(f"Поле {field_name} містить недопустимі символи.")
    return value.strip()

def load_config():
    config = ConfigParser()
    config.read("config.txt")

    try:
        fb_profile_path = os.path.join(os.getcwd(), "fb_profile")
        kijiji_profile_path = os.path.join(os.getcwd(), "fb_profile")  # Так само, як Facebook

        return {
            "telegram_token": config.get("settings", "telegram_token").strip(),
            "telegram_chat_id": config.get("settings", "telegram_chat_id").strip(),
            "city_fb": validate_text_field(config.get("filters", "city_fb"), "city_fb").lower(),
            "city_kijiji": validate_text_field(config.get("filters", "city_kijiji"), "city_kijiji").lower(),
            "make": validate_text_field(config.get("filters", "make"), "make").lower(),
            "model": validate_text_field(config.get("filters", "model"), "model").lower(),
            "year_min": int(config.get("filters", "year_min")),
            "year_max": int(config.get("filters", "year_max")),
            "price_min": int(config.get("filters", "price_min")),
            "price_max": int(config.get("filters", "price_max")),
            "facebook_days_filter": int(config.get("filters", "facebook_days_filter")),
            "max_post_age_hours": int(config.get("filters", "max_post_age_hours")),
            "interval": int(config.get("settings", "interval")),
            "facebook_profile_path": fb_profile_path,
            "facebook_profile_name": "Default",
            "kijiji_profile_path": kijiji_profile_path,
            "kijiji_profile_name": "Default",
        }
    except Exception as e:
        print(f"❌ Помилка при читанні конфігурації: {e}")
        raise

def main():
    last_mtime = None
    config = None
    last_run_time = 0

    while True:
        try:
            current_mtime = os.path.getmtime("config.txt")
            if last_mtime != current_mtime:
                print("🔄 Оновлення конфігурації...")
                config = load_config()
                print(f"🛠 Завантажена конфігурація: interval={config['interval']} хв, make={config['make']}, model={config['model']}")
                last_mtime = current_mtime
                # Скидаємо таймер, щоб новий інтервал почався після оновлення
                last_run_time = 0

            if config is None:
                time.sleep(1)
                continue

            if config['interval'] <= 0:
                print("❌ Інтервал оновлення має бути більше 0!")
                time.sleep(10)
                continue

            current_time = time.time()
            if current_time - last_run_time >= config['interval'] * 60:
                print("🚀 Запускаємо парсинг...")

                try:
                    fb_ads = parse_facebook(config)
                except Exception as e:
                    print(f"⚠️ Помилка при парсингу Facebook: {e}")
                    fb_ads = []

                try:
                    kj_ads = parse_kijiji(config)
                except Exception as e:
                    print(f"⚠️ Помилка при парсингу Kijiji: {e}")
                    kj_ads = []

                all_ads = fb_ads + kj_ads
                if all_ads:
                    filename = save_ads(all_ads, config["make"], config["model"], config["city_fb"], f"Last {config['facebook_days_filter']} days")
                    send_file_to_telegram(filename, config["telegram_token"], config["telegram_chat_id"])
                else:
                    print("🔔 Нових оголошень не знайдено.")

                save_seen()
                last_run_time = current_time

        except Exception as e:
            print(f"❌ Помилка в основному циклі: {e}")

        time.sleep(1)


if __name__ == "__main__":
    main()
