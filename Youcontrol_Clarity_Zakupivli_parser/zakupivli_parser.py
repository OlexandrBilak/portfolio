import re
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def parse_popup(driver, wait, url):
    driver.get(url)

    try:
        button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.zkb-company-bubble")))
        button.click()
        time.sleep(1)  # час на анімацію

        popup = driver.execute_script("""
            const btn = arguments[0];
            let popup = btn.parentElement.querySelector('div.zkb-popup__body');
            if (!popup) {
                popup = document.querySelector('div.zkb-popup__body');
            }
            return popup;
        """, button)

        if popup:
            driver.execute_script("""
                const popup = arguments[0];
                popup.style.visibility = 'visible';
                popup.style.opacity = '1';
                popup.style.height = 'auto';
                popup.style.width = 'auto';
                popup.style.maxHeight = 'none';
                popup.style.maxWidth = 'none';
                popup.style.overflow = 'visible';
            """, popup)
            time.sleep(1)

            text = driver.execute_script("return arguments[0].innerText", popup)

            name_match = re.search(r"Ім'я:\s*(.+)", text)
            email_match = re.search(r"E-mail:\s*([\w\.-]+@[\w\.-]+)", text)
            phone_match = re.search(r"Телефон:\s*([\+\d\s\-]+)", text)

            name = name_match.group(1).strip() if name_match else ""
            email = email_match.group(1).strip() if email_match else ""
            phone = phone_match.group(1).strip() if phone_match else ""

            # Перевірка довжини даних
            name = name if len(name) > 1 else "Немає"
            email = email if len(email) > 1 else "Немає"
            phone = phone if len(phone) > 1 else "Немає"

            return name, email, phone
        else:
            return "Немає", "Немає", "Немає"

    except Exception as e:
        print(f"Помилка при парсингу: {e}")
        return "Немає", "Немає", "Немає"

def main():
    # Параметри
    input_file = "output.xlsx"  # Заміни на свій файл
    sheet_name = "Sheet1"      # Заміни на свій лист
    edrpou_col = "E"           # Колонка з ЄДРПОУ
    start_row = 2              # З якої строки починаємо (1 - заголовок)
    max_to_process = 177        # Скільки оголошень обробляти

    # Колонки для запису даних (права від ЄДРПОУ)
    name_col = "I"
    email_col = "J"
    phone_col = "K"

    # Відкриваємо Excel
    wb = load_workbook(input_file)
    ws = wb[sheet_name]

    # Ініціалізація selenium
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 3)

    processed_count = 0

    for row in range(start_row, ws.max_row + 1):
        if processed_count >= max_to_process:
            break

        edrpou = ws[f"{edrpou_col}{row}"].value
        if not edrpou:
            continue

        # Перевіряємо чи є вже дані в колонках
        name_cell = ws[f"{name_col}{row}"].value
        email_cell = ws[f"{email_col}{row}"].value
        phone_cell = ws[f"{phone_col}{row}"].value

        if all([name_cell, email_cell, phone_cell]):
            print(f"Рядок {row} пропускаємо, дані вже є")
            continue

        print(f"Обробка рядка {row}, ЄДРПОУ: {edrpou}")

        # Формуємо URL для запиту (з твоїм параметром)
        url = f"https://zakupivli.pro/gov/tenders?q={edrpou}"

        name, email, phone = parse_popup(driver, wait, url)

        # Записуємо у Excel
        ws[f"{name_col}{row}"] = name
        ws[f"{email_col}{row}"] = email
        ws[f"{phone_col}{row}"] = phone

        processed_count += 1

    # Зберігаємо результат
    wb.save("output.xlsx")

    driver.quit()
    print("Готово!")

if __name__ == "__main__":
    main()
